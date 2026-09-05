"""
NBSC PRIME-HRM Intelligence Hub — Employee Import Services
Parses Excel and CSV rosters to import employees into MongoDB.
"""
import re
import openpyxl
from datetime import datetime
from .models import Employee


def generate_unique_employee_id(department: str) -> str:
    """Generates next available employee ID, e.g. NBSC-2026-0042."""
    year = datetime.utcnow().year
    count = Employee.objects.count() + 1
    return f"NBSC-{year}-{count:04d}"


def import_employees_from_excel(file_obj) -> dict:
    """
    Parses an uploaded Excel workbook (.xlsx) and imports employees.
    Supports standard columns:
    [Employee ID, Last Name, First Name, Middle Name, Email, Department, Position, Category, Status, Daily Rate]
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    imported = 0
    skipped = 0
    errors = []

    # Detect header row
    header_row_idx = 1
    headers = {}
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[r]]
        if any('name' in v or 'email' in v or 'position' in v or 'employee' in v for v in row_vals):
            header_row_idx = r
            for idx, val in enumerate(row_vals):
                if val:
                    headers[val] = idx
            break

    # If standard headers not found, look for CAR sheet format
    if 'car-' in ws.title.lower() or not any('email' in k for k in headers):
        return import_from_car_sheet(ws)

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = ws[row_idx]
        if not any(c.value for c in row):
            continue

        def get_col(*aliases):
            for a in aliases:
                for h_name, col_idx in headers.items():
                    if a in h_name:
                        val = row[col_idx].value
                        return str(val).strip() if val is not None else ''
            return ''

        first_name = get_col('first', 'given')
        last_name = get_col('last', 'surname')
        full_name = get_col('name')

        if not first_name and full_name:
            # Parse full name
            parts = full_name.split(',', 1)
            if len(parts) == 2:
                last_name = parts[0].strip()
                first_name = parts[1].strip()
            else:
                parts = full_name.split()
                if len(parts) >= 2:
                    first_name = parts[0]
                    last_name = " ".join(parts[1:])
                else:
                    first_name = full_name
                    last_name = 'Staff'

        if not first_name or not last_name:
            continue

        email = get_col('email')
        if not email:
            slug = re.sub(r'[^a-z0-9]', '', f"{first_name.lower()}.{last_name.lower()}")
            email = f"{slug}@nbsc.edu.ph"

        dept = get_col('department', 'office', 'inst') or 'DGEC'
        dept_clean = 'DGEC'
        for d in ['DGEC', 'IBM', 'ICS', 'ITE', 'ADMIN', 'FIN', 'REG']:
            if d.lower() in dept.lower():
                dept_clean = d
                break

        pos = get_col('position', 'designation', 'rank') or 'Instructor I'
        cat = 'TEACHING' if any(w in pos.lower() for w in ['instructor', 'prof', 'faculty', 'teacher']) else 'NON_TEACHING'
        status = get_col('status', 'employment') or 'COS'

        emp_id = get_col('id', 'employee id', 'emp_id') or generate_unique_employee_id(dept_clean)

        existing = Employee.objects(email=email).first()
        if existing:
            skipped += 1
            continue

        try:
            emp = Employee(
                employee_id=emp_id,
                first_name=first_name,
                last_name=last_name,
                email=email,
                department=dept_clean,
                position=pos,
                category=cat,
                employment_status='COS' if 'cos' in status.lower() or 'contract' in status.lower() else 'PERMANENT',
                daily_rate=1325.68 if cat == 'TEACHING' else 850.0,
                monthly_salary=29165.0 if cat == 'TEACHING' else 18700.0,
                date_hired=datetime.utcnow()
            )
            emp.save()
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    return {
        'imported': imported,
        'skipped': skipped,
        'errors': errors
    }


def import_from_car_sheet(ws) -> dict:
    """Specialized parser for NBSC Comparative Assessment Report sheets."""
    imported = 0
    skipped = 0
    errors = []

    pos_title = 'Instructor I'
    dept_code = 'DGEC'

    # Read sheet metadata
    for r in range(1, 14):
        vals = [c.value for c in ws[r] if c.value is not None]
        for v in vals:
            s = str(v)
            if 'Vacant Position' in s:
                pass
            if 'General Education' in s or 'DGEC' in s:
                dept_code = 'DGEC'
            elif 'Business' in s or 'IBM' in s:
                dept_code = 'IBM'
            elif 'Computer' in s or 'ICS' in s:
                dept_code = 'ICS'
            elif 'Teacher Education' in s or 'ITE' in s:
                dept_code = 'ITE'

    # Rows 15 onwards contain candidate names
    for r in range(15, ws.max_row + 1):
        c2 = ws.cell(row=r, column=2).value
        if not c2 or 'nothing follows' in str(c2).lower():
            break

        name_str = str(c2).strip()
        if not name_str or name_str == '-' or len(name_str) < 3:
            continue

        # Split name into first and last
        parts = name_str.split()
        if len(parts) >= 2:
            first_name = parts[0]
            last_name = " ".join(parts[1:])
        else:
            first_name = name_str
            last_name = "Candidate"

        clean_slug = re.sub(r'[^a-z0-9]', '', f"{first_name.lower()}.{parts[-1].lower()}")
        email = f"{clean_slug}@nbsc.edu.ph"

        if Employee.objects(email=email).first():
            skipped += 1
            continue

        try:
            emp = Employee(
                employee_id=generate_unique_employee_id(dept_code),
                first_name=first_name,
                last_name=last_name,
                email=email,
                department=dept_code,
                position=pos_title,
                category='TEACHING',
                employment_status='COS',
                daily_rate=1325.68,
                monthly_salary=29165.0,
                date_hired=datetime.utcnow()
            )
            emp.save()
            imported += 1
        except Exception as ex:
            errors.append(f"Row {r}: {str(ex)}")

    return {
        'imported': imported,
        'skipped': skipped,
        'errors': errors
    }
